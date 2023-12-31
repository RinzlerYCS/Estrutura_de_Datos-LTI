import sys
import sqlite3
from sqlite3 import Error
import datetime as dt
import re
import csv
import openpyxl
import pandas as pd
from openpyxl import Workbook

fecha_actual = dt.date.today()
try:
    with sqlite3.connect ('TallerMecanico.db') as conn:
        mi_cursor=conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Cliente (clave INTEGER PRIMARY KEY, nombre TEXT NOT NULL, rfc TEXT NOT NULL, correo TEXT NOT NULL);")
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Servicios  (clave_servicio INTEGER PRIMARY KEY, nombre_servicio TEXT NOT NULL, precio INTEGER NOT NULL );")
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Notas (folio INTEGER PRIMARY KEY, fecha timestamp ,activo INTEGER NOT NULL, clave INTEGER NOT NULL, FOREIGN KEY (clave) REFERENCES Cliente (clave));")
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Detalles (folio_detalles INTEGER PRIMARY KEY, clave_servicio INTEGER ,  folio INTEGER ,FOREIGN KEY (clave_servicio) REFERENCES Servicios (clave_servicio),FOREIGN KEY (folio) REFERENCES Notas (folio));")
except Error as e:
    print (e)
except Exception:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
finally:
    conn.close()
while True:
    try:
        print('')
        print('1. Notas\n2. Clientes\n3. Servicios\n4. Salir')
        opcion = int(input('Selecciona una opción: '))
    except ValueError:
        print("Por favor, ingrese un número válido.")
        continue
    match opcion:
        case 1: #notas
            while True:
                clave_c=0
                print('1. Registrar una nota\n2. Cancelar una nota\n3. Recuperar una nota\n4. Consultas y reportes\n5. Volver al menú principal')
                opcion2 = int(input('Selecciona una opción en el submenú: '))
                match opcion2:
                    case 1:
                        try:
                          with sqlite3.connect("TallerMecanico.db") as conn:
                            my_cursor=conn.cursor()
                            my_cursor.execute("SELECT clave,nombre FROM Cliente;")
                            querie=my_cursor.fetchall()
                            print("Opción 1 Registrar una nota\n")
                            print("Claves\tNombre")
                            print("-" * 30)
                            for clave,nombre in querie:#mostrar clave | nombree
                              print(f"{clave:^6}|\t{nombre}")
                            print("-" * 30)
                        except Error as e:
                          print(e)
                        finally:
                          conn.close()
                        
                        try:
                          with sqlite3.connect('TallerMecanico.db') as conn:
                            while True:
                              my_cursor=conn.cursor()
                              clave_de_cliente=int(input("Eliga una de las claves mostradas: "))
                              diccionario_cliente={"clave":clave_de_cliente}
                              my_cursor.execute("SELECT * FROM Cliente  WHERE clave=:clave ",diccionario_cliente)
                              querie=my_cursor.fetchall()

                              if querie:
                                print(f"Si existe este cliente con clave: {clave_de_cliente}")
                                clave_c=clave_de_cliente
                                break
                              else:
                                print("Este cliente no existe o se encuentra cancelado")
                                continue

                        except Error as e:
                          print(e)
                        except ValueError:
                          print("El dato debe de ser numerico")
                          continue
                        finally:
                          conn.close()
                        
                        while True:##VALIDACION DE FECHAS
                          try:
                             fecha_registro_nota=input("\nIngrese su fecha con el formato (dd/mm/aaaa): ")#SE PIDE LA FECHA DEL USUARIO
                             if fecha_registro_nota.strip()=="":
                               print("EL DATO NO DEBE DE OMITIRSE")
                               continue
                             fecha_procesada = dt.datetime.strptime(fecha_registro_nota, "%d/%m/%Y").date()
                          except ValueError:
                                print("El dato ingresado no cumple con el formato (dd/mm/yy)")
                                continue
                          if fecha_procesada>=fecha_actual:            #SE VERIFICA QUE LA FECHA NO SEA posterior A LA ACTUAL
                               print("La fecha que usted ingreso es invalida, su fecha no debe ser posterior a la actual")
                               continue
                          else:
                                break

                        try:
                          with sqlite3.connect('TallerMecanico.db') as conn:
                            my_cursor=conn.cursor()

                            a={"fecha":fecha_procesada,
                                "activo":1,
                                "clave":clave_c}
                            my_cursor.execute("INSERT INTO Notas (fecha,activo,clave) VALUES (:fecha,:activo,:clave)",a)
                        except Error as e:#SE CREA LA NOTA, SOLO FALTAN LOS DETALLES
                            print(e)
                        finally:
                            conn.close()

                        try:
                            with sqlite3.connect('TallerMecanico.db') as conn:
                              my_cursor=conn.cursor()
                              my_cursor.execute("SELECT clave_servicio, nombre_servicio, precio FROM Servicios;")
                              querie=my_cursor.fetchall()
                              print("Eliga alguno de los siguientes servicios")
                              print("Clave\t\tNombre\t\t\tPrecio")
                              print("-" * 45)
                              for clave,nombre,preciod in querie:#mostrar clave | nombree
                                print(f"{clave:^6}|\t{nombre}\t|\t{preciod}")
                              print("-" * 45)
                        except Error as e:
                          print(e)
                        finally:
                          conn.close()



                        while True:
                          try:
                            with sqlite3.connect('TallerMecanico.db') as conn:
                              my_cursor=conn.cursor()
                              clave_servicio=int(input("Ingrese alguna de las claves mostradas previamente: "))##FALTA VALIDAR LA LLAVE DEL SERVICIO
                              validacion_servicio_diccionario={"clave_servicio":clave_servicio}
                              my_cursor.execute("SELECT * FROM Servicios WHERE clave_servicio=:clave_servicio",validacion_servicio_diccionario)
                              validacion_servicio=my_cursor.fetchall()

                              if validacion_servicio:
                                print("Si existe este servicio")
                              else:
                                print("Este servicio no existe")
                                continue
                          except Error as e:
                            print(e)
                          finally:
                            conn.close()


                          try:
                            with sqlite3.connect('TallerMecanico.db') as conn:
                              my_cursor=conn.cursor()
                              my_cursor.execute("SELECT folio FROM Notas ORDER BY folio DESC LIMIT 1")
                              conexx_N_D=my_cursor.fetchone()
                              b={"clave_servicio":clave_servicio,
                              "folio":conexx_N_D[0]}
                              
                              my_cursor.execute("INSERT INTO Detalles (clave_servicio,folio)VALUES (:clave_servicio, :folio)",b)

                              otro_servicio=int(input("Desea elegir otro servicio? \n1-Si 2-No"))
                              if otro_servicio ==1:
                                continue
                              elif otro_servicio==2:
                                break
                              else:
                                print("Ingrese 1-Si   2-No")
                                continue
                          except Error as e:
                            print(e)
                          finally:
                            conn.close()


                        print("SI LLEGO HASTA AQUI SE AGREGARA A LA BD")##FALTA AGREGAR LA INFO A LA TABLA DETALLES



                    case 2:
                        print("Opción 2 Cancelar una nota")
                        folio_cancelar_global=0
                        while True:

                          try:
                            with sqlite3.connect('TallerMecanico.db') as conn:
                              my_cursor=conn.cursor()
                              folio_n_cancelar=int(input("Ingrese su folio de la nota a cancelar: "))
                              
                              en_sentencia8={"folio":folio_n_cancelar}
                              my_cursor.execute("SELECT * FROM Notas WHERE folio=:folio AND activo='1'",en_sentencia8)
                              mostrar_c=my_cursor.fetchall()
                              for a,b,c,d in mostrar_c:
                                print(a,b,c,d)

                              if mostrar_c :
                                print("EXISTE")
                                #Si la nota existe, se desplegarán los 
                                #datos de la nota indicada y todo el detalle que la conforma
                                folio_cancelar_global=folio_n_cancelar
                              else:
                                print("La nota indicada no esta en el sistema o se encuentra cancelada")
                                break

                          except ValueError:
                            print('el dato debe de ser numerico')
                            continue
                          except Error as e:
                            print(e)
                          finally:
                            conn.close()

                          try:
                            with sqlite3.connect('TallerMecanico.db') as conn:
                              my_cursor=conn.cursor()
                              
                              mostrar_nota_c_detalles={"folio":folio_cancelar_global}
                              my_cursor.execute("SELECT N.folio,N.fecha,C.nombre,C.rfc, \
                              C.correo,SUM(S.precio)AS 'monto a pagar',S.nombre_servicio, N.activo \
                              FROM Notas N INNER JOIN CLIENTE C ON N.clave=C.clave\
                              INNER JOIN Detalles D ON D.folio=N.folio \
                              INNER JOIN Servicios S ON S.clave_servicio=D.clave_servicio \
                              WHERE N.folio=:folio AND N.activo=1",mostrar_nota_c_detalles)

                              mostrar_c=my_cursor.fetchall()
    
                              print("folio\tfecha\tnombre\trfc\tcorreo\tmonto\tservicios\tactivo")
                              print("-" * 50)
                                #se muestra con detalles
                              for foliof, fechaf, nombref, correof, rfcf, montof, serviciosf, activof in mostrar_c:
                                  print(foliof, fechaf, nombref, rfcf, correof, montof, serviciosf, activof)
                          except Error as e:
                              print(e)
                          finally:
                            conn.close()

                          while True:
                              try:
                                confirmacion_cancelar=int(input("Desea cancelar la nota?\n1-Si 2-No: "))
                                with sqlite3.connect("TallerMecanico.db") as conn:
                                  my_cursor=conn.cursor()
                                  if confirmacion_cancelar ==1:
                                    print("SE VA A CANCELAR LA NOTA")
                                    cancelar_nota_dicc={"folio":folio_cancelar_global}
                                    my_cursor.execute("UPDATE Notas SET activo=0 WHERE folio=(:folio);",cancelar_nota_dicc)
                                    break
                                  elif confirmacion_cancelar==2:
                                    print("No se cancelara la nota")
                                    break
                                  elif confirmacion_cancelar>2 or confirmacion_cancelar<1:
                                    print("Debe ingresar la opcion 1 o 2")
                                    continue
                              except Error as e:
                                print(e)
                              except ValueError:
                                print("El dato debe de ser numerico")
                              finally:
                                conn.close()
                                
                          break
                        break
                         


                    case 3:
                        print("Opción 3 Recuperar una nota")
                        
                        while True:

                            
                            folio_cancelados=[]
                            print("Notas Canceladas")
                            try:
                              with sqlite3.connect('TallerMecanico.db') as conn:
                                my_cursor=conn.cursor()

                                my_cursor.execute("SELECT T.folio, T.nombre, T.rfc, T.correo, T.fecha, T.[Monto a pagar]\
                                                  FROM (\
                                                      SELECT N.folio, C.nombre, C.rfc, C.correo, N.fecha, SUM(S.precio) AS 'Monto a pagar'\
                                                      FROM Notas N\
                                                      INNER JOIN Cliente C ON N.clave = C.clave\
                                                      INNER JOIN Detalles D ON D.folio = N.folio\
                                                      INNER JOIN Servicios S ON S.clave_servicio = D.clave_servicio\
                                                      WHERE activo = 0\
                                                      GROUP BY N.folio, C.nombre, C.rfc, C.correo, N.fecha\
                                                        ) AS T")
                                mostrar_c=my_cursor.fetchall()##SE MUESTRAN LAS CANCELADAS SIN SUS DETALLES
                                for nfolio, cnombre, crfc, ccorreo, nfecha,monto_pagar in mostrar_c:
                                  print(nfolio, cnombre, crfc, ccorreo, nfecha,monto_pagar)
                                                 
                                  folio_cancelados.append(mostrar_c[0][0])                 
                            except Error as e:
                              print(e)


                            finally:
                              conn.close()

                            try:
                              decision_recuperar_nota = int(input("Desea recuperar alguna de las siguientes notas\n1-Si  2-No: "))  
                              if decision_recuperar_nota == 1:
                                  pass
                              elif decision_recuperar_nota == 2: 
                                  break
                              else:
                                  print("Opción no válida. Por favor, elija 1 o 2.")
                            except ValueError:
                                print("Error: Ingrese un número entero válido (1 o 2).")
                            
                            try:
                              folio_n_cancelar=int(input("Ingrese su folio a recuperar: "))
                            except ValueError:#falta validar que exista
                                print('el dato debe de ser numerico')
                                continue
                            try:
                              with sqlite3.connect('TallerMecanico.db') as conn:
                                my_cursor=conn.cursor()
                                #SE MUESTRAN LOS DETALLES DE LA NOTA
                                consulta_f={"folio":folio_n_cancelar}
                                my_cursor.execute("SELECT S.nombre_servicio, S.precio FROM Notas N INNER JOIN \
                                Detalles D ON N.folio=D.folio INNER JOIN Servicios S ON D.clave_servicio=S.clave_servicio WHERE N.folio=:folio ",consulta_f)
                                muestra_d_detalles=my_cursor.fetchall()
                                for name,price in muestra_d_detalles:
                                  print(name,price)
                            except Error as e:
                              print(e)
                            finally:
                              conn.close()
                    

                            while True:
                              try:
                                confirmar_recuperar=int(input("Desea recuperar la nota? \n1-Si 2-No: "))
                              except ValueError:
                                print("El dato debe de ser numerico ")
                                continue

                              if confirmar_recuperar==2:
                                break
                              elif confirmar_recuperar==1:
                                try:
                                  with sqlite3.connect('TallerMecanico.db') as conn:
                                    my_cursor=conn.cursor()
                                    print("se recuperara la noTa")
                                    recuperar_n={"folio":folio_n_cancelar}
                                    my_cursor.execute("UPDATE Notas SET activo=1 WHERE folio=(:folio);",recuperar_n)
                                    break
                                except Error as e:
                                  print(e)
                                finally:
                                  conn.close() 
                                                      
                            break
                        break



                    case 4:
                        print("Opción 4 Consultas y reportes")  #consultas y reportes
                        while True:
                            print('1. Consulta por periodo\n2. Consulta por folio\n3. Volver al menú principal')
                            opcion3 = int(input('Selecciona una opción en el submenú: '))
                            match opcion3:
                                case 1:
                                    print("Opción 1 Consulta por periodo")

                                    while True:   #CONSULTA POR PERIODO
                                      try:

                                          fecha_inicio = input("Ingrese la fecha inicial del período, en formato dd/mm/aaaa:\n")
                                          if fecha_inicio.strip() == "":
                                              print("Se tomara la fecha predeterminada como 01/01/2000")
                                              fecha_inicio ="1/1/2000"
                                          fecha_inicio = dt.datetime.strptime(fecha_inicio, "%d/%m/%Y").date()
                                          if fecha_inicio>fecha_actual:            #SE VERIFICA QUE LA FECHA NO SEA posterior A LA ACTUAL
                                              print("La fecha que usted ingreso es invalida, su fecha no debe ser posterior a la actual")
                                              continue


                                      except ValueError:
                                          print("El dato ingresado no cumple con el patrón (dd/mm/aaaa)")
                                          continue

                                      try:
                                          fecha_final = input("Ingrese la fecha final del período, en formato dd/mm/aaaa:\n")
                                          if fecha_final.strip() == "":###validar que sea posterior
                                              print("No se ingreso nada, por lo tanto la fecha predetermianda es la fecha actual")
                                              fecha_final =dt.date.today()
                                          else:
                                              fecha_final = dt.datetime.strptime(fecha_final, "%d/%m/%Y").date()
                                      except ValueError:
                                          print("El dato ingresado no cumple con el patrón (dd/mm/aaaa)")
                                          continue

                                      #usar diccionario

                                      try:
                                        with sqlite3.connect('TallerMecanico.db') as conn:
                                          my_cursor=conn.cursor()
                                          diccion_fecha={"fecha_inicial":fecha_inicio,
                                                          "fecha_final":fecha_final
                                                          }
                                          my_cursor.execute("SELECT	N.folio, N.fecha, C.nombre, C.rfc, C.correo, SUM(S.precio) as 'Monto a pagar' \
                                            FROM Notas N INNER JOIN Cliente C ON N.clave=C.clave INNER \
                                            JOIN Detalles D ON D.folio = N.folio INNER JOIN Servicios S ON D.clave_servicio=S.clave_servicio \
                                            WHERE N.fecha BETWEEN :fecha_inicial AND :fecha_final AND N.activo=1\
                                            GROUP BY N.folio, N.fecha, C.nombre, C.rfc, C.correo;",diccion_fecha)
                                          notas_fechas=my_cursor.fetchall()
                                          if notas_fechas:
                                            pass
                                          else:
                                            print("No hay notas emitidas para dicho periodo")
                                            break




                                          print("folio\tfecha\t\tnombre\t\trfc\t\tcorreo\t\tmonto a pagar")
                                          print("-" * 100)
                                      
                                          for a,b,c,d,e,f in notas_fechas:##SE MUESTRAN LAS NOTAS SIN SU DETALLE
                                            print(f"{a}|\t{b}|\t{c}|  {d}|\t{e}|\t{f}|")

                                          df = pd.DataFrame(notas_fechas, columns=["folio", "fecha", "nombre", "rfc", "correo", "monto a pagar"])



                                          pasar_a_csv_excel=int(input("Desea pasar esta informacion a \n1-.csv  2-Excel  3-Volver al menú:  "))
                                          if pasar_a_csv_excel==1:
                                            print("Se pasara a .csv")
                                            df.to_csv(f"Reporte_por_periodo_{fecha_inicio}_{fecha_final}", index=False)
                                            break
                                          elif pasar_a_csv_excel==2:
                                            print("Se pasara a Excel")
                                            df.to_excel(f"Reporte_por_periodo_{fecha_inicio}_{fecha_final}.xlsx", index=False)
                                            break
                                          elif pasar_a_csv_excel==3:
                                            break

                                            
                                      except Error as e:
                                        print(e)
                                      finally :
                                        conn.close()

                                case 2:
                                    print("Opción 2 Consulta por folio")

                                    try:
                                      with sqlite3.connect('TallerMecanico.db') as conn:
                                        my_cursor=conn.cursor()
                                        my_cursor.execute("SELECT N.folio, N.fecha, C.nombre FROM Notas N INNER \
                                        JOIN Cliente C ON N.clave=C.clave WHERE activo='1'ORDER BY folio;")
                                        folio_consulta_cursor=my_cursor.fetchall()

                                        print("Folio\tFecha\t\t\tNombre")#SE  MUESTRAN LAS NOTAS
                                        print("-" * 45)
                                        for folio,fecha,nombre in folio_consulta_cursor:
                                          print(f"{folio:^6}|\t{fecha}|\t{nombre}|")
                                        print("-" * 45)
                                        
                                    except Error as e:
                                      print(e)
                                    finally:
                                      conn.close()
                                    while True:

                                      try: 
                                        with sqlite3.connect('TallerMecanico.db') as conn:
                                          my_cursor=conn.cursor()
                                          folio_a_consultar=int(input("Ingrese su folio a consultar: "))
                                          nominal_a_consultar={"folio":folio_a_consultar}
                                          my_cursor.execute("SELECT * FROM Notas WHERE activo=1 AND folio=:folio",nominal_a_consultar)
                                          validacion_consulta_f=my_cursor.fetchall()
                                          if validacion_consulta_f:
                                            print(f"existe el folio: {folio_a_consultar}")
                                          else:
                                            print("Este folio no existe o está cancelado")
                                            continue
                                      except Error as e:
                                        print(e)
                                      finally:
                                        conn.close()

                                      try:
                                        with sqlite3.connect('TallerMecanico.db') as conn:
                                          my_cursor=conn.cursor()

                                          en_sentencia7={"folio":folio_a_consultar}
                                          my_cursor.execute("SELECT	N.folio, N.fecha, C.nombre,C.nombre, C.rfc, C.correo, \
                                          SUM(S.precio)as 'Monto a pagar' FROM Notas N INNER JOIN Cliente C ON N.clave=C.clave \
                                          INNER JOIN Detalles D ON D.folio = N.folio INNER JOIN Servicios S ON D.clave_servicio=S.clave_servicio\
                                            WHERE N.folio=:folio",en_sentencia7)
                                          
                                          mostrar_c=my_cursor.fetchall()
                                          if mostrar_c:
                                            for folio, fecha, nombre,nombre, rfc, correo,monto in mostrar_c:
                                              print(folio, fecha, nombre,nombre, rfc, correo,monto)
                                            break
                                          else:
                                            print("Folio cancelado o no existente")
                                            continue
                                      except ValueError:
                                        print("Error al ingresar el dato")
                                      except Error as e:
                                        print(e)
                                      finally:
                                        conn.close()
                                    ##SI EL FOLIO NO EXISTE


                                case 3:

                                    break
                                case _:
                                    print("Opción no válida")
                    case 5:
                      break
        case 2: #clientes
            while True:
                print('')
                print('1. Agregar un cliente\n2. Consultas y reportes\n3. Volver al menú principal')
                try:
                    opcion4 = int(input('Selecciona una opción en el submenú: '))
                except ValueError:
                    print("Por favor, ingrese un número válido.")
                    continue 
                match opcion4:
                    case 1:
                        while True:
                            nombre_cliente=input("Ingrese su nombre: ")
                            if nombre_cliente.strip()=="":
                                print("El nombre no tiene que estar en blanco")
                            else:
                                break
                        while True:
                            patron_rfc = (r'^[A-Z&Ñ]{3,4}\d{6}[A-V1-9][A-Z1-9]\d{1}$')
                            rfc_usuario = input("Introduce tu RFC: ")
                            if rfc_usuario.strip()=="":
                                print("El RFC no tiene que estar en blanco")
                                continue
                            if re.match(patron_rfc, rfc_usuario):
                                break
                            else:
                                print("RFC inválido, vuelva a ingresarlo")
                        while True:
                            patron_correo = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                            correo = input("Ingrese una dirección de correo electrónico: ")
                            if correo.strip()=="":
                                print("El correo no tiene que estar en blanco")
                            if re.match(patron_correo, correo):
                                break
                            else:
                                print("La dirección de correo electrónico no es válida.")
                        try:
                            with sqlite3.connect("TallerMecanico.db") as conn:
                                mi_cursor = conn.cursor() 
                                valores = (nombre_cliente, rfc_usuario, correo )#No se incluye dato para la PK
                                mi_cursor.execute("INSERT INTO Cliente (nombre, rfc, correo ) VALUES(?,?,?)", valores)
                        except Error as e:
                            print (e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()  
                    case 2:
                        while True:
                            print('')
                            print('1. Listados de clientes registrados\n2. Busqueda por clave\n3. Busqueda por nombrel\n4. Volver al menú principal')
                            try:
                                opcion5 = int(input('Selecciona una opción en el submenú: '))
                            except ValueError:
                                print("Por favor, ingrese un número válido.")
                                continue
                            match opcion5:
                                case 1:
                                    while True:
                                        print('')
                                        print('1. Ordenar por clave\n2. Ordenar por nombre\n3. Volver al menú principal')
                                        try:
                                            opcion6 = int(input('Selecciona una opción en el submenú: '))
                                        except ValueError:
                                            print("Por favor, ingrese un número válido.")
                                            continue
                                        match opcion6:
                                            case 1:
                                                try:
                                                    with sqlite3.connect("TallerMecanico.db") as conn:
                                                        mi_cursor = conn.cursor()
                                                        mi_cursor.execute("SELECT * FROM Cliente ORDER BY clave")
                                                        clientesactivos= mi_cursor.fetchall()
                                                        if clientesactivos:
                                                            print("Lista de clientes activos")
                                                            for fila in clientesactivos:
                                                                print(f"Clave: {fila[0]} \t Nombre: {fila[1]} \t RFC: {fila[2]} \t Correo: {fila[3]}")
                                                            while True:
                                                                print('Donde deseas guardar el arvhivo\n 1. CSV \n2. Excel\n3. Salir')
                                                                try:
                                                                   opcionarchivo1= int(input('Selecciona una opción en el submenú: '))
                                                                except ValueError:
                                                                    print("Por favor, ingrese solo números .")
                                                                    continue
                                                                match opcionarchivo1:
                                                                    case 1:
                                                                        nombre_archivocsv = f"ReporteClientesActivosPorClave_{fecha_actual}.csv"
                                                                        with open(nombre_archivocsv, mode='w', newline='') as archivo_csv:
                                                                            escritor_csv = csv.writer(archivo_csv)
                                                                            escritor_csv.writerow(['Clave', 'Nombre', 'RFC', 'Correo'])
                                                                            for fila in clientesactivos:
                                                                                escritor_csv.writerow([fila[0], fila[1], fila[2], fila[3]])
                                                                        print(f"Los datos se han guardado en el archivo CSV '{nombre_archivocsv}'.")
                                                                        break
                                                                    case 2:
                                                                        nombre_archivo_xlsx = f"ReporteClientesActivosPorClave_{fecha_actual}.xlsx"
                                                                        workbook = Workbook()
                                                                        sheet = workbook.active
                                                                        sheet.append(['Clave', 'Nombre', 'RFC', 'Correo'])
                                                                        for fila in clientesactivos:
                                                                            sheet.append([fila[0], fila[1], fila[2], fila[3]])
                                                                        workbook.save(nombre_archivo_xlsx)
                                                                        print(f"Los datos se han guardado en el archivo Excel '{nombre_archivo_xlsx}'.")
                                                                        break

                                                                    case 3:
                                                                        break
                                                                    case _:
                                                                        print("Opción no válida")  
                                                        else:
                                                            print(f"No se encuentran clientes")
                                                except Error as e:
                                                    print (e)
                                                except Exception:
                                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                                finally:
                                                    conn.close()
                                            case 2:
                                                try:
                                                    with sqlite3.connect("TallerMecanico.db") as conn:
                                                        mi_cursor = conn.cursor()
                                                        mi_cursor.execute("SELECT * FROM Cliente ORDER BY nombre")
                                                        clientesactivos= mi_cursor.fetchall()
                                                        if clientesactivos:
                                                            print("Lista de clientes activos")
                                                            for fila in clientesactivos:
                                                                print(f"Nombre: {fila[1]} \t Clave: {fila[0]} \t RFC: {fila[2]} \t Correo: {fila[3]}")
                                                        while True:
                                                           print('Donde deseas guardar el arvhivo\n 1. CSV \n2. Excel\n3. Salir')
                                                           try:
                                                               opcionarchivo1= int(input('Selecciona una opción en el submenú: '))
                                                           except ValueError:
                                                                print("Por favor, ingrese solo números .")
                                                                continue
                                                           match opcionarchivo1:
                                                                case 1:
                                                                    nombre_archivocsv = f"ReporteClientesActivosPorNombre_{fecha_actual}.csv"
                                                                    with open(nombre_archivocsv, mode='w', newline='') as archivo_csv:
                                                                        escritor_csv = csv.writer(archivo_csv)
                                                                        escritor_csv.writerow(['Clave', 'Nombre', 'RFC', 'Correo'])
                                                                        for fila in clientesactivos:
                                                                            escritor_csv.writerow([fila[0], fila[1], fila[2], fila[3]])
                                                                    print(f"Los datos se han guardado en el archivo CSV '{nombre_archivocsv}'.")
                                                                    break 
                                                                case 2:
                                                                    nombre_archivo_xlsx = f"ReporteClientesActivosPorNombre_{fecha_actual}.xlsx"
                                                                    workbook = Workbook()
                                                                    sheet = workbook.active
                                                                    sheet.append(['Clave', 'Nombre', 'RFC', 'Correo'])
                                                                    for fila in clientesactivos:
                                                                        sheet.append([fila[0], fila[1], fila[2], fila[3]])
                                                                    workbook.save(nombre_archivo_xlsx)
                                                                    print(f"Los datos se han guardado en el archivo Excel '{nombre_archivo_xlsx}'.")
                                                                    break
                                                                case 3:
                                                                    break
                                                                case _:
                                                                    print("Opción no válida")    
                                                        else:
                                                            print(f"No se encuentran clientes")
                                                except Error as e:
                                                    print (e)
                                                except Exception:
                                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                                finally:
                                                    conn.close()
                                            case 3:
                                                break
                                            case _:
                                                print("Opción no válida")
                                case 2:
                                    while True:
                                        try:
                                            clave_busqueda = int(input("Ingrese la clave del cliente a buscar: "))
                                        except ValueError:
                                            print("Por favor, ingrese solo números enteros como clave del cliente.")
                                            continue 
                                        try:
                                            with sqlite3.connect("TallerMecanico.db") as conn:
                                                mi_cursor = conn.cursor()
                                                mi_cursor.execute("SELECT * FROM Cliente WHERE clave=?", (clave_busqueda,))
                                                clavecliente= mi_cursor.fetchone()
                                                if clavecliente:
                                                    print("Cliente encontado")
                                                    print(f"Clave: {clavecliente[0]}, Nombre: {clavecliente[1]}, RFC: {clavecliente[2]}, Correo: {clavecliente[3]}")
                                                else:
                                                    print(f"No se encuentran clientes")
                                        except Error as e:
                                            print (e)
                                        except Exception:
                                             print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                        finally:
                                            conn.close()
                                        break
                                case 3:
                                    nombre_busqueda = input("Ingrese el nombre del cliente a buscar: ")
                                    try:
                                        with sqlite3.connect("TallerMecanico.db") as conn:
                                            mi_cursor = conn.cursor()
                                            mi_cursor.execute("SELECT * FROM Cliente WHERE nombre=?", (nombre_busqueda,))
                                            nombrecliente= mi_cursor.fetchone()
                                            if nombrecliente:
                                                print("Cliente encontado")
                                                print(f"Clave: {nombrecliente[0]}, Nombre: {nombrecliente[1]}, RFC: {nombrecliente[2]}, Correo: {nombrecliente[3]}")
                                            else:
                                                print(f"No se encuentran clientes")
                                    except Error as e:
                                        print (e)
                                    except Exception:
                                         print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                    finally:
                                        conn.close()
                                                          
                                case 4:
                                    break
                                case _:
                                    print("Opción no válida")
                    case 3:
                        break  
                    case _:
                        print("Opción no válida")
        case 3: #Servicios
            while True:
                print('')
                print('1. Agregar un servicio\n2. Consultas y reportes\n3. Volver al menú principal')
                try:
                    opcion7 = int(input('Selecciona una opción en el submenú: '))
                except ValueError:
                    print("Por favor, ingrese un número válido.")
                    continue
                match opcion7:
                    case 1:
                        while True:
                            nombre_servicio=input("Ingrese el nombre del servicio: ").lower()
                            if nombre_servicio.strip()=="":
                                print("El nombre no tiene que estar en blanco")
                            else:
                                break
                        while True:
                            try:
                              precio_servicio=float(input("Que costo deseas que tenga el servicio"))
                              if precio_servicio<= (0):
                                print("El precio tiene que ser mayor que $0.00")
                                continue
                              else:
                                break
                            except ValueError:
                                print("El valor ingresado tiene que tener un numero")
                        try:
                            with sqlite3.connect("TallerMecanico.db") as conn:
                                mi_cursor = conn.cursor()
                                valores_servicios=(nombre_servicio,precio_servicio)
                                mi_cursor.execute("INSERT INTO Servicios(nombre_servicio,precio)VALUES(?,?)",valores_servicios)
                                print("Su servicio se agrego correctamente")
                        except Error as e:
                            print (e)
                        except:
                            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                        finally:
                            conn.close()
                    case 2:
                        while True:
                            print('')
                            print('1. Busqueda por clave del servicio\n2. Busqueda por nombre del servicio\n3. Listado de servicios\n4. Volver al menú principal')
                            try:
                                opcion8 = int(input('Selecciona una opción en el submenú: '))
                            except ValueError:
                                print("Por favor, ingrese solo números enteros como clave del cliente.")
                                continue
                            match opcion8:
                                case 1:
                                    try:
                                      with sqlite3.connect("TallerMecanico.db") as conn:
                                          mi_cursor = conn.cursor()
                                          mi_cursor.execute("SELECT clave_servicio, nombre_servicio FROM Servicios")
                                          filas = mi_cursor.fetchall()
                                          if filas:
                                              print("Lista de servicios")
                                              for r in filas:
                                                  print(f"Clave: {r[0]} \t Nombre: {r[1]} ")
                                          else:
                                              print(f"No se encuentran clientes")
                                          
                                      while True:
                                          servicio_clave = input("Ingresa la clave del servicio que deseas consultar: ")
                                          if servicio_clave.strip()=="":
                                            print("El campo no puede estar vacio")
                                            continue
                                          elif not servicio_clave.isdigit():
                                            print("El campo debe tener un numero")
                                            continue
                                          else:
                                                                         
                                            mi_cursor.execute("SELECT*FROM Servicios WHERE clave_servicio=?",(servicio_clave))
                                            servicio_encontrado=mi_cursor.fetchone()
                                            if servicio_encontrado:
                                              print("Servicio Encontrado") 
                                              print(f"Clave: {servicio_encontrado[0]} \t Nombre: {servicio_encontrado[1]} \t Precio: {servicio_encontrado[2]} ")
                                        
                                              break
                                            else:
                                                print("El servicio no existe")
                                                continue 
                                                break
                                    except sqlite3.Error as e:
                                        print(e)
                                    except Exception as e:
                                        print(f"Se produjo el siguiente error: {e}, vuelve a intentarlo")
                                case 2:
                                    while True:
                                        servicio_busqueda=input("Ingresa el servicio que deseas consultar : ").lower() #para las mayusculas y minusculas
                                        if servicio_busqueda.strip()=="":
                                            print("El campo debe no puede quedar vacio")
                                        else:
                                           break
                                    try:
                                      with sqlite3.connect("TallerMecanico.db")as conn:
                                        mi_cursor = conn.cursor()
                                        mi_cursor.execute("SELECT*FROM Servicios WHERE nombre_servicio=?", (servicio_busqueda,))
                                        nombre_servicio=mi_cursor.fetchone()
                                        if nombre_servicio[1]== servicio_busqueda:
                                          print(f"Servicio Encontrado \n Nombre: {nombre_servicio[1]} \t Precio;  ${nombre_servicio[2]}")     
                                    except Error as e:
                                        print (e)
                                    except Exception:
                                          print(f"El servicio que buscas no existe")
                                    finally:
                                        conn.close()
                                        break
                                case 3:
                                    while True:
                                        print('1. Ordenar por clave\n2. Ordenar por nombre del servicio\n3. Volver al menú principal')
                                        try:
                                            opcion9 = int(input('Selecciona una opción en el submenú: '))
                                        except ValueError:
                                            print("Por favor, ingrese solo números ")
                                            continue
                                        match opcion9:
                                            case 1:
                                                try:
                                                   with sqlite3.connect("TallerMecanico.db") as conn:
                                                       mi_cursor = conn.cursor()
                                                       mi_cursor.execute("SELECT * FROM Servicios ORDER BY clave_servicio")
                                                       claves_activas = mi_cursor.fetchall()
                                                       if claves_activas:
                                                           for fila in claves_activas:                                                  
                                                                 print(f"{fila[0]:^6}|\t{fila[1]},|\t{fila[2]}")
                                                       while True:
                                                           print('Donde deseas guardar el arvhivo\n 1. CSV \n2. Excel\n3. Salir')
                                                           try:
                                                               opcionarchivo = int(input('Selecciona una opción en el submenú: '))
                                                           except ValueError:
                                                                print("Por favor, ingrese solo números .")
                                                                continue
                                                           match opcionarchivo:
                                                                case 1:
                                                                    nombre_archivocsv= f"ReporteServiciosPorClave_{fecha_actual}.csv"
                                                                    with open(nombre_archivocsv, mode='w', newline='') as archivo_csv:
                                                                        escritor_csv = csv.writer(archivo_csv)
                                                                        escritor_csv.writerow(['Clave', 'Nombre del Servicio', 'Precio'])
                                                                        for fila in claves_activas:
                                                                            escritor_csv.writerow(fila)
                                                                        break 
                                                                case 2:
                                                                    nombre_archivo_xlsx = f"ReporteServiciosPorClave_{fecha_actual}.xlsx"
                                                                    workbook = Workbook()
                                                                    sheet = workbook.active
                                                                    sheet.append(['Clave', 'Nombre del Servicio', 'Precio'])
                                                                    for fila in claves_activas:
                                                                        sheet.append(fila)
                                                                    workbook.save(nombre_archivo_xlsx)
                                                                    print(f"Los datos se han guardado en el archivo Excel '{nombre_archivo_xlsx}'.")

                                                                case 3:
                                                                    break
                                                                case _:
                                                                    print("Opción no válida")
                                                                    
                                                except Error as e:
                                                    print (e)
                                                finally:
                                                     conn.close()
                                            case 2:
                                                try:
                                                    with sqlite3.connect("TallerMecanico.db") as conn:
                                                        mi_cursor = conn.cursor()
                                                        mi_cursor.execute("SELECT * FROM Servicios ORDER BY nombre_servicio")
                                                        claves_activas = mi_cursor.fetchall()
                                                        if claves_activas:
                                                            print(f"Nombre\tClaves\tPrecio")
                                                            for fila in claves_activas:
                                                                print(f"{fila[1]:^6}|\t{fila[0]},|\t{fila[2]}")
                                                        while True:
                                                            print('Donde deseas guardar el arvhivo\n 1. CSV \n2. Excel\n3. Salir')
                                                            try:
                                                                opcionarchivo = int(input('Selecciona una opción en el submenú: '))
                                                            except ValueError:
                                                                print("Por favor, ingrese solo números .")
                                                                continue
                                                            match opcionarchivo:
                                                                case 1:
                                                                    nombre_archivocsv = f"ReporteServiciosPorNombre_{fecha_actual}.csv"
                                                                    with open(nombre_archivocsv, mode='w', newline='') as archivo_csv:
                                                                        escritor_csv = csv.writer(archivo_csv)
                                                                        escritor_csv.writerow(['Clave', 'Nombre del Servicio', 'Precio'])
                                                                        escritor_csv.writerow([nombre_servicio[0], nombre_servicio[1], nombre_servicio[2]])
                                                                    print(f"Los datos se han guardado en el archivo CSV '{nombre_archivocsv}'.")
                                                                    break
                                                                case 2:
                                                                    nombre_archivo_xlsx = f"ReporteServiciosPorNombre_{fecha_actual}.xlsx"
                                                                    workbook = Workbook()
                                                                    sheet = workbook.active
                                                                    sheet.append(['Clave', 'Nombre del Servicio', 'Precio'])
                                                                    for fila in claves_activas:
                                                                        sheet.append([fila[0], fila[1], fila[2]])
                                                                    workbook.save(nombre_archivo_xlsx)
                                                                    print(f"Los datos se han guardado en el archivo Excel '{nombre_archivo_xlsx}'.")
                                                                    break
                                                                case 3:
                                                                    break
                                                                case _:
                                                                    print("Opción no válida")  
                                                except Error as e:
                                                    print (e)     
                                                finally:
                                                    conn.close()     
                                            case 3:
                                                break
                                            case _:
                                                print("Opción no válida")
                                case 4:
                                    break
                                case _:
                                    print("Opción no válida")
                    case 3:
                        break
                    case _:
                        print("Opción no válida")
        case 4:
            break
        case _:
            print("Opción no válida")
